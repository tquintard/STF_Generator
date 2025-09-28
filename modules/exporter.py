from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, Union, List
import os
import re
import shutil
import unicodedata

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from config.settings import OUTPUT_DIR, log, BASE_DIR


TEMPLATE_MAP_STATE_KEY = "_stf_template_map"
_TEMPLATE_MAP_CACHE: Optional[Dict[str, str]] = None


def _get_template_map() -> Dict[str, str]:
    global _TEMPLATE_MAP_CACHE
    session_map = None
    try:
        session_map = st.session_state.get(TEMPLATE_MAP_STATE_KEY)
    except Exception:
        session_map = None
    if session_map is not None:
        _TEMPLATE_MAP_CACHE = dict(session_map)
        return dict(session_map)

    if _TEMPLATE_MAP_CACHE is None:
        _TEMPLATE_MAP_CACHE = _discover_template_files()
    template_map = dict(_TEMPLATE_MAP_CACHE)
    try:
        st.session_state[TEMPLATE_MAP_STATE_KEY] = template_map
    except Exception:
        pass
    return template_map


def _resolve_template_for_sgapp(sgapp_value: Union[str, float, None]) -> Optional[str]:
    raw = str(sgapp_value).strip() if sgapp_value is not None else ""
    if not raw or raw.lower() in {"nan", "none"}:
        return None
    key = raw.upper()
    template_map = _get_template_map()
    return template_map.get(key)


def _normalize_header(name: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(name))
    return "".join(ch for ch in normalized if not unicodedata.combining(ch)).lower()


def _ensure_output_dir():
    try:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
    except Exception as exc:
        log(f"Failed to ensure output dir {OUTPUT_DIR}: {exc}", level="ERROR")


def _infer_attr_type(attr: str, mda_df: Optional[pd.DataFrame]) -> str:
    if mda_df is None or mda_df.empty or "MUDU" not in mda_df.columns:
        return "str"
    row = mda_df.loc[mda_df["MUDU"] == attr]
    if row.empty:
        return "str"
    normalized = {_normalize_header(col): col for col in mda_df.columns}
    for candidate in ("type", "attrtype", "datatype", "attributetype", "columntype"):
        col = normalized.get(candidate)
        if not col:
            continue
        value = str(row.iloc[0][col]).strip().lower()
        if any(token in value for token in ("int", "integer")):
            return "int"
        if any(token in value for token in ("float", "double", "decimal", "number", "numeric", "real")):
            return "float"
        break
    return "str"


def _to_number(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in {"nan", "none", "null", "-"}:
        return None
    s = s.replace("\u00A0", " ").replace(" ", "").replace("'", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _excel_safe(name: str) -> str:
    s = str(name).strip()
    s = re.sub(r"[\s\-]+", "_", s)
    s = re.sub(r"[^A-Za-z0-9_]", "", s)
    if s and s[0].isdigit():
        s = f"_{s}"
    return s


def _safe_dirname(name: str, default: str) -> str:
    value = str(name).strip() if name is not None else ""
    if not value:
        value = default
    sanitized = re.sub(r'[^0-9A-Za-z_-]+', '_', value)
    return sanitized or default


def _iter_defined_names(wb):
    dn = wb.defined_names
    try:
        for obj in dn.definedName:
            yield obj
        return
    except AttributeError:
        pass
    try:
        for obj in dn.values():
            if isinstance(obj, (list, tuple)):
                for item in obj:
                    yield item
            else:
                yield obj
        return
    except Exception:
        pass
    try:
        for key in dn.keys():
            obj = dn.get(key)
            if isinstance(obj, (list, tuple)):
                for item in obj:
                    yield item
            elif obj is not None:
                yield obj
    except Exception:
        return


def _write_by_named_ranges(wb, values: Dict[str, str], mda_df: Optional[pd.DataFrame]):
    names_map: Dict[str, tuple] = {}
    for dn in _iter_defined_names(wb):
        if getattr(dn, "type", None) != "RANGE":
            continue
        dests = list(getattr(dn, "destinations", []) or [])
        if not dests:
            continue
        sheet_name, ref = dests[0]
        cell = ref.split(":", 1)[0]
        key = dn.name.strip()
        variants = {key, key.lower(), _excel_safe(key),
                    _excel_safe(key).lower()}
        for variant in variants:
            if variant and variant not in names_map:
                names_map[variant] = (sheet_name, cell)

    for key, value in values.items():
        variants = [key, key.lower(), _excel_safe(key),
                    _excel_safe(key).lower()]
        target = None
        for variant in variants:
            target = names_map.get(variant)
            if target:
                break
        if not target:
            continue
        sheet_name, cell_ref = target
        if sheet_name not in wb.sheetnames:
            continue
        cell = wb[sheet_name][cell_ref]
        cell_type = _infer_attr_type(key, mda_df)
        if cell_type in ("int", "float"):
            number = _to_number(value)
            cell.value = number if number is not None else value
        else:
            cell.value = value


def _get_defined_base_names(wb) -> set:
    names = set()
    for dn in _iter_defined_names(wb):
        name = getattr(dn, "name", "") or ""
        lower = name.lower()
        if lower.startswith("_xlnm.") or lower in {"print_area", "print_titles"}:
            continue
        names.add(name)
    return names


def _fill_workbook_with_values(wb, values: Dict[str, str], mda_df: Optional[pd.DataFrame]):
    _write_by_named_ranges(wb, values, mda_df)


def export_stf_from_row(selected: Union[pd.DataFrame, Dict], template_path: Optional[str] = None) -> str:
    if isinstance(selected, pd.DataFrame):
        row = selected.iloc[0]
        ecs = str(row.get("RepeFonct", "")).strip()
        values = {str(k): ("" if pd.isna(v) else str(v))
                  for k, v in row.items()}
    elif isinstance(selected, dict):
        ecs = str(selected.get("RepeFonct", "")).strip()
        values = {str(k): ("" if v is None else str(v))
                  for k, v in selected.items()}
    else:
        raise TypeError(
            "selected must be a pandas DataFrame (1 row) or a dict of values")

    sgapp_value = values.get("SGApp")
    sg_value = sgapp_value.split("-")[0]
    sgapp_raw = str(sgapp_value).strip() if sgapp_value is not None else ""
    if template_path is None:
        template_path = _resolve_template_for_sgapp(sgapp_value)
        if template_path is None:
            template_dir = Path(BASE_DIR) / "template"
            display_sgapp = sgapp_raw or "unknown"
            log(f"Template not found for SGApp={display_sgapp}", level="ERROR")
            raise FileNotFoundError(
                f"Template not found for SGApp '{display_sgapp}'")

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    safe_ecs = "".join(ch for ch in ecs if ch.isalnum()
                       or ch in ("-", "_")) or "ECS"

    elemsys_dir = _safe_dirname(values.get("ElemSys"), "ElemSys")
    sgapp_dir = _safe_dirname(sg_value, "SGApp")
    target_dir = os.path.join(OUTPUT_DIR, elemsys_dir, sgapp_dir)
    os.makedirs(target_dir, exist_ok=True)

    filename = f"{sgapp_value}_{safe_ecs}.xlsx"
    output_path = os.path.join(target_dir, filename)
    counter = 1
    while os.path.exists(output_path):
        filename = f"{safe_ecs}_{counter}.xlsx"
        output_path = os.path.join(target_dir, filename)
        counter += 1

    shutil.copyfile(template_path, output_path)

    wb = load_workbook(output_path)
    mda_df = st.session_state.get("mda_df")

    if isinstance(mda_df, pd.DataFrame) and not mda_df.empty and 'MUDU' in mda_df.columns:
        mda_keys = set(mda_df['MUDU'].astype(str))
        values = {k: v for k, v in values.items() if k in mda_keys}

    _fill_workbook_with_values(wb, values, mda_df)
    wb.save(output_path)
    log(f"STF exported for {ecs} -> {output_path}")

    try:
        if isinstance(mda_df, pd.DataFrame) and not mda_df.empty and 'MUDU' in mda_df.columns:
            defined_names = _get_defined_base_names(wb)
            mda_variants = {s.lower() for s in mda_df['MUDU'].astype(str)}
            mda_variants |= {_excel_safe(s).lower()
                             for s in mda_df['MUDU'].astype(str)}
            extras = [n for n in defined_names if n.lower()
                      not in mda_variants]
            if extras:
                st.info(
                    f"Defined names in template not present in MDA: {', '.join(sorted(extras))}"
                )
    except Exception:
        pass
    return output_path


def _apply_values_to_sheet(ws, cell_map: Dict[str, str], values: Dict[str, str], mda_df: Optional[pd.DataFrame]):
    lowered = {str(k).lower(): v for k, v in values.items()}
    for key, cell_ref in cell_map.items():
        cell = ws[cell_ref]
        value = values.get(key)
        if value is None:
            value = values.get(key.lower())
        if value is None:
            value = lowered.get(key.lower())
        if value is None:
            safe = _excel_safe(key)
            value = values.get(safe)
            if value is None:
                value = lowered.get(safe.lower())
        if value is None:
            value = ""
        cell_type = _infer_attr_type(key, mda_df)
        if cell_type in ("int", "float"):
            number = _to_number(value)
            cell.value = number if number is not None else value
        else:
            cell.value = value


def _collect_named_ranges_by_sheet(wb) -> Dict[str, Dict[str, str]]:
    sheet_map: Dict[str, Dict[str, str]] = {}
    for dn in _iter_defined_names(wb):
        if getattr(dn, "type", None) != "RANGE":
            continue
        destinations = list(getattr(dn, "destinations", []) or [])
        if not destinations:
            continue
        sheet_name, ref = destinations[0]
        cell = ref.split(":", 1)[0]
        sheet_map.setdefault(sheet_name, {})
        variants = {
            dn.name,
            dn.name.lower(),
            _excel_safe(dn.name),
            _excel_safe(dn.name).lower(),
        }
        for variant in variants:
            if variant:
                sheet_map[sheet_name][variant] = cell
    return sheet_map


def _discover_template_files() -> Dict[str, str]:
    template_dir = Path(BASE_DIR) / "template"
    if not template_dir.exists():
        return {}
    pattern = re.compile(r"^(SG\d{2}-\d{2}-\d)_\w{3}", re.IGNORECASE)
    mapping: Dict[str, str] = {}
    for path in template_dir.glob("*.xlsx"):
        match = pattern.match(path.stem)
        if match:
            mapping[match.group(1).upper()] = str(path)
    return mapping


def _safe_sheet_title(base: str, existing: List[str]) -> str:
    sanitized = "".join(ch for ch in base if ch.isalnum()
                        or ch in (" ", "-", "_"))
    sanitized = sanitized.strip() or "Sheet"
    if len(sanitized) > 31:
        sanitized = sanitized[:31]
    candidate = sanitized
    suffix = 1
    while candidate in existing or len(candidate) == 0:
        suffix += 1
        trimmed = sanitized[:28] if len(sanitized) > 28 else sanitized
        candidate = f"{trimmed}_{suffix}"
        if len(candidate) > 31:
            candidate = candidate[:31]
    return candidate


def export_stf_batch(data_df: pd.DataFrame) -> List[str]:
    if data_df is None or data_df.empty:
        return []

    template_map = _get_template_map()
    if not template_map:
        log("No STF templates discovered in template directory", level="ERROR")
        return []

    df = data_df.copy()
    if not {"SGApp", "ElemSys"}.issubset(df.columns):
        log("SGApp/ElemSys columns missing; aborting batch export", level="ERROR")
        return []

    df.sort_values(by=["SGApp", "ElemSys", "RepeFonct"],
                   inplace=True, kind="mergesort")

    exported_paths: List[str] = []
    mda_df = st.session_state.get("mda_df")

    for (elemsys, sgapp), group in df.groupby(["ElemSys", "SGApp"], dropna=False):
        sgapp_raw = str(sgapp).strip()
        sgapp_key = sgapp_raw.upper()
        sg_key = sgapp_key.split("-")[0]
        template_path = template_map.get(sgapp_key)
        if not template_path:
            log(f"Template not found for SGApp={sgapp_raw}", level="WARNING")
            continue

        group_sorted = group.sort_values(by="RepeFonct", kind="mergesort")
        if group_sorted.empty:
            continue

        _ensure_output_dir()
        elemsys_key = str(elemsys).strip() or "ElemSys"
        safe_elemsys = "".join(
            ch for ch in elemsys_key if ch.isalnum() or ch in ("-", "_")) or "ElemSys"
        elemsys_dir = _safe_dirname(elemsys_key, "ElemSys")
        sgapp_dir = _safe_dirname(sg_key, "SGApp")
        target_dir = os.path.join(OUTPUT_DIR, elemsys_dir, sgapp_dir)
        os.makedirs(target_dir, exist_ok=True)
        filename = f"STF_{safe_elemsys}_{sgapp_key}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        output_path = os.path.join(target_dir, filename)
        try:
            shutil.copyfile(template_path, output_path)
        except Exception as exc:
            log(
                f"Failed to copy template '{template_path}' to '{output_path}': {exc}", level="ERROR")
            continue

        try:
            agg_wb = load_workbook(output_path, data_only=False)
        except Exception as exc:
            log(f"Failed to load workbook '{output_path}': {exc}",
                level="ERROR")
            continue

        try:
            template_wb = load_workbook(template_path, data_only=False)
        except Exception as exc:
            log(f"Failed to reload template '{template_path}': {exc}", level="ERROR")
            agg_wb.close()
            continue

        template_sheet_names = template_wb.sheetnames
        range_map = _collect_named_ranges_by_sheet(template_wb)

        existing_titles: List[str] = []

        for _, row in group_sorted.iterrows():
            row_values = {str(k): ("" if pd.isna(v) else str(v))
                          for k, v in row.items()}
            if isinstance(mda_df, pd.DataFrame) and not mda_df.empty and 'MUDU' in mda_df.columns:
                mda_keys = set(mda_df['MUDU'].astype(str))
                row_values = {k: v for k, v in row_values.items()
                              if k in mda_keys}

            base_name = str(row.get("RepeFonct", "ECS")).strip() or "ECS"

            for sheet_name in template_sheet_names:
                if sheet_name not in agg_wb.sheetnames:
                    continue
                template_ws = agg_wb[sheet_name]
                new_ws = agg_wb.copy_worksheet(template_ws)
                new_title = _safe_sheet_title(
                    base_name if len(
                        template_sheet_names) == 1 else f"{base_name}_{sheet_name}",
                    existing_titles,
                )
                new_ws.title = new_title
                existing_titles.append(new_title)

                cell_map = range_map.get(sheet_name, {})
                _apply_values_to_sheet(new_ws, cell_map, row_values, mda_df)

        for sheet_name in template_sheet_names:
            if sheet_name in agg_wb.sheetnames:
                agg_wb.remove(agg_wb[sheet_name])

        try:
            agg_wb.save(output_path)
            agg_wb.close()
            exported_paths.append(output_path)
            log(f"Exported STF batch -> {output_path} | sheets={len(existing_titles)}")
        except Exception as exc:
            log(f"Failed to save workbook '{output_path}': {exc}",
                level="ERROR")

    return exported_paths
